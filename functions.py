import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
import io

def initial_df_to_base(path='базаа.xlsx'):
    initial_df = pd.read_excel(path)
    initial_df['Аналоги_split'] = initial_df['Аналоги'].apply(lambda x: x.split('\n'))
    exploded_df = initial_df.explode('Аналоги_split')
    exploded_df['Аналоги_split'] = exploded_df['Аналоги_split'].apply(lambda x: x.split(' - ')[1].strip(' ,.-'))
    df_base = exploded_df[['Каталожный номер', 'Аналоги_split']].astype(str).reset_index(drop=True)

    return df_base


def request_df_to_request_list(path='запрос.xlsx'):
    df_request = pd.read_excel(path)
    request_numbers_list = df_request['Каталожный номер'].astype(str).to_list()

    return request_numbers_list


def num_finder(df_base, request_numbers_list):
    nums_list = []
    for num in request_numbers_list:
        for i in df_base.index:
            cat_num = df_base.loc[i, 'Каталожный номер']
            analog_num = df_base.loc[i, 'Аналоги_split']
            if num == cat_num or num == analog_num:
                nums_list.append(num)

    return nums_list


def request_handler(nums_list, path='базаа.xlsx'):
    rows_list = []
    initial_df = pd.read_excel(path)
    for i in initial_df.index:
        for num in nums_list:
            if (num in str(initial_df.loc[i, 'Каталожный номер'])) or (num in str(initial_df.loc[i, 'Аналоги'])):
                rows_list.append(initial_df.loc[[i]])
    return_df = pd.concat(rows_list)
    return_df = return_df.drop_duplicates(keep='first')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        return_df.to_excel(writer, index=False)
    output.seek(0)

    workbook = openpyxl.load_workbook(output)
    sheet = workbook['Sheet1']
    fill = PatternFill(start_color="ebf051", end_color="ebf051", fill_type="solid")

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)
        for num in nums_list:
            if num in str(cell.value):
                cell.fill = fill



    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=8)
        for num in nums_list:
            if num in str(cell.value):
                cell.fill = fill


    return workbook


